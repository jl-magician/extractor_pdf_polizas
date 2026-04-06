"""Export module for writing polizas data to Excel and CSV formats.

Implements EXP-01 (xlsx multi-sheet), EXP-02 (CSV flat), EXP-05 (type-correct cells).
"""
from __future__ import annotations

import csv
from decimal import Decimal
from pathlib import Path
from typing import Sequence

from policy_extractor.storage.models import Asegurado, Cobertura, Poliza


class ExportError(Exception):
    """Raised when the output file cannot be written (e.g. file open in Excel)."""


# ---------------------------------------------------------------------------
# Formatting constants
# ---------------------------------------------------------------------------

DATE_FMT = "DD/MM/YYYY"
MONEY_FMT = "#,##0.00"

# Columns exported for each table (excludes provenance and evaluation fields)
POLIZA_COLUMNS = [
    "numero_poliza",
    "aseguradora",
    "tipo_seguro",
    "fecha_emision",
    "inicio_vigencia",
    "fin_vigencia",
    "nombre_contratante",
    "nombre_agente",
    "prima_neta",
    "derecho_poliza",
    "recargo",
    "descuento",
    "iva",
    "otros_cargos",
    "prima_total",
    "primer_pago",
    "pago_subsecuente",
    "moneda",
    "forma_pago",
    "frecuencia_pago",
]

ASEGURADO_COLUMNS = [
    "numero_poliza",  # FK for VLOOKUP — injected from parent Poliza
    "tipo",
    "nombre_descripcion",
    "fecha_nacimiento",
    "rfc",
    "curp",
    "direccion",
    "parentesco",
]

COBERTURA_COLUMNS = [
    "numero_poliza",  # FK for VLOOKUP — injected from parent Poliza
    "nombre_cobertura",
    "suma_asegurada",
    "deducible",
    "moneda",
]

DATE_COLS = {"fecha_emision", "inicio_vigencia", "fin_vigencia", "fecha_nacimiento"}
MONEY_COLS = {
    "prima_total", "prima_neta", "derecho_poliza", "recargo", "descuento",
    "iva", "otros_cargos", "primer_pago", "pago_subsecuente", "suma_asegurada", "deducible",
}


# ---------------------------------------------------------------------------
# Internal helpers
# ---------------------------------------------------------------------------


def _cell_value(val):
    """Coerce Python types for openpyxl compatibility.

    - Decimal → float  (openpyxl writes Decimal as string otherwise)
    - date / datetime  → pass through  (openpyxl handles natively)
    - None             → None
    - everything else  → as-is
    """
    if isinstance(val, Decimal):
        return float(val)
    return val


def _collect_extra_keys(items, attr: str = "campos_adicionales") -> list[str]:
    """Two-pass: collect all unique keys from campos_adicionales across all items.

    Strips the 'confianza' key (internal storage metadata, not user data).
    Preserves insertion order using dict as ordered set.
    """
    seen: dict[str, None] = {}
    for item in items:
        extras: dict = getattr(item, attr) or {}
        for k in extras:
            if k != "confianza":
                seen[k] = None
    return list(seen)


def _apply_formats(ws, header_row: list[str]) -> None:
    """Apply number_format to date and monetary columns after data is written.

    Must be called after all rows are appended — setting format per-cell after
    ws.append() is the only reliable approach in openpyxl 3.x.
    """
    from openpyxl.utils import get_column_letter

    for col_idx, col_name in enumerate(header_row, start=1):
        col_letter = get_column_letter(col_idx)
        if col_name in DATE_COLS:
            fmt = DATE_FMT
        elif col_name in MONEY_COLS:
            fmt = MONEY_FMT
        else:
            continue
        # Skip header row (index 0 in the column tuple)
        for cell in ws[col_letter][1:]:
            cell.number_format = fmt


def _finalize_sheet(ws) -> None:
    """Apply auto-filter and freeze header row.

    Auto-filter is only set when there is at least one data row (max_row > 1)
    to avoid ws.dimensions returning 'A1:A1' on empty sheets.
    """
    if ws.max_row > 1:
        ws.auto_filter.ref = ws.dimensions
    ws.freeze_panes = "A2"


def _write_sheet(
    ws,
    header: list[str],
    rows: list[list],
) -> None:
    """Generic sheet writer: append header then data rows, then format and finalize."""
    ws.append(header)
    for row in rows:
        ws.append(row)
    _apply_formats(ws, header)
    _finalize_sheet(ws)


# ---------------------------------------------------------------------------
# Row builders
# ---------------------------------------------------------------------------


def _poliza_rows(
    polizas: Sequence[Poliza],
    extra_keys: list[str],
) -> list[list]:
    """Build flat row list for the polizas sheet."""
    rows = []
    for poliza in polizas:
        base = [_cell_value(getattr(poliza, col, None)) for col in POLIZA_COLUMNS]
        extras_dict: dict = dict(poliza.campos_adicionales or {})
        extras_dict.pop("confianza", None)
        extra_vals = [_cell_value(extras_dict.get(k, "")) for k in extra_keys]
        rows.append(base + extra_vals)
    return rows


def _asegurado_rows(
    polizas: Sequence[Poliza],
    extra_keys: list[str],
) -> list[list]:
    """Build flat row list for the asegurados sheet.

    Injects numero_poliza from the parent poliza for VLOOKUP support.
    """
    # ASEGURADO_COLUMNS[0] == "numero_poliza" (injected), rest are ORM attrs
    non_fk_cols = ASEGURADO_COLUMNS[1:]  # skip "numero_poliza"
    rows = []
    for poliza in polizas:
        for aseg in poliza.asegurados:
            base = [poliza.numero_poliza] + [
                _cell_value(getattr(aseg, col, None)) for col in non_fk_cols
            ]
            extras_dict: dict = dict(aseg.campos_adicionales or {})
            extras_dict.pop("confianza", None)
            extra_vals = [_cell_value(extras_dict.get(k, "")) for k in extra_keys]
            rows.append(base + extra_vals)
    return rows


def _cobertura_rows(
    polizas: Sequence[Poliza],
    extra_keys: list[str],
) -> list[list]:
    """Build flat row list for the coberturas sheet.

    Injects numero_poliza from the parent poliza for VLOOKUP support.
    """
    non_fk_cols = COBERTURA_COLUMNS[1:]  # skip "numero_poliza"
    rows = []
    for poliza in polizas:
        for cob in poliza.coberturas:
            base = [poliza.numero_poliza] + [
                _cell_value(getattr(cob, col, None)) for col in non_fk_cols
            ]
            extras_dict: dict = dict(cob.campos_adicionales or {})
            extras_dict.pop("confianza", None)
            extra_vals = [_cell_value(extras_dict.get(k, "")) for k in extra_keys]
            rows.append(base + extra_vals)
    return rows


# ---------------------------------------------------------------------------
# Public API
# ---------------------------------------------------------------------------


def export_xlsx(polizas: Sequence[Poliza], output_path: Path) -> int:
    """Write a 3-sheet Excel workbook to output_path.

    Sheets: polizas, asegurados, coberturas.
    Monetary cells are numeric floats; date cells are Excel date types with DD/MM/YYYY format.
    campos_adicionales JSON keys are expanded into individual columns (confianza stripped).
    Auto-filter and frozen header row applied to every sheet.

    Args:
        polizas: Sequence of Poliza ORM objects with relationships eagerly loaded.
        output_path: Destination .xlsx file path.

    Returns:
        Number of polizas exported.

    Raises:
        ExportError: If the file cannot be written (e.g. open in another application).
    """
    # Lazy import — keeps openpyxl out of module-load path for fast CLI startup
    from openpyxl import Workbook

    # Collect union-of-all extra keys for each sheet (confianza already stripped)
    poliza_extra_keys = _collect_extra_keys(polizas, "campos_adicionales")

    all_asegurados: list[Asegurado] = [
        aseg for p in polizas for aseg in p.asegurados
    ]
    aseg_extra_keys = _collect_extra_keys(all_asegurados, "campos_adicionales")

    all_coberturas: list[Cobertura] = [
        cob for p in polizas for cob in p.coberturas
    ]
    cob_extra_keys = _collect_extra_keys(all_coberturas, "campos_adicionales")

    wb = Workbook()

    # --- Sheet 1: polizas ---
    ws_polizas = wb.active
    ws_polizas.title = "polizas"
    _write_sheet(
        ws_polizas,
        POLIZA_COLUMNS + poliza_extra_keys,
        _poliza_rows(polizas, poliza_extra_keys),
    )

    # --- Sheet 2: asegurados ---
    ws_asegurados = wb.create_sheet("asegurados")
    _write_sheet(
        ws_asegurados,
        ASEGURADO_COLUMNS + aseg_extra_keys,
        _asegurado_rows(polizas, aseg_extra_keys),
    )

    # --- Sheet 3: coberturas ---
    ws_coberturas = wb.create_sheet("coberturas")
    _write_sheet(
        ws_coberturas,
        COBERTURA_COLUMNS + cob_extra_keys,
        _cobertura_rows(polizas, cob_extra_keys),
    )

    try:
        wb.save(output_path)
    except (PermissionError, OSError) as exc:
        raise ExportError(
            f"Cannot write to {output_path} — close the file if it is open in another application"
        ) from exc

    return len(list(polizas))


def export_csv(polizas: Sequence[Poliza], output_path: Path) -> int:
    """Write a flat UTF-8 BOM CSV of polizas to output_path.

    Only the polizas table is exported (use export_xlsx for asegurados/coberturas).
    campos_adicionales JSON keys are expanded into individual columns (confianza stripped).

    Args:
        polizas: Sequence of Poliza ORM objects.
        output_path: Destination .csv file path.

    Returns:
        Number of polizas exported.

    Raises:
        ExportError: If the file cannot be written.
    """
    polizas_list = list(polizas)
    extra_keys = _collect_extra_keys(polizas_list, "campos_adicionales")
    fieldnames = POLIZA_COLUMNS + extra_keys

    try:
        with open(output_path, "w", encoding="utf-8-sig", newline="") as f:
            writer = csv.DictWriter(f, fieldnames=fieldnames, extrasaction="ignore")
            writer.writeheader()
            for poliza in polizas_list:
                row: dict = {
                    col: _cell_value(getattr(poliza, col, None))
                    for col in POLIZA_COLUMNS
                }
                extras_dict: dict = dict(poliza.campos_adicionales or {})
                extras_dict.pop("confianza", None)
                for k in extra_keys:
                    row[k] = _cell_value(extras_dict.get(k, ""))
                writer.writerow(row)
    except (PermissionError, OSError) as exc:
        raise ExportError(
            f"Cannot write to {output_path} — close the file if it is open in another application"
        ) from exc

    return len(polizas_list)
