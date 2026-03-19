"""Unit tests for policy_extractor.export module (EXP-01, EXP-02, EXP-05).

Also contains CLI integration tests for EXP-03, EXP-04 (format routing, Spanish filters).
"""
from __future__ import annotations

import csv
import io
import json
from datetime import date
from decimal import Decimal
from pathlib import Path
from unittest.mock import patch

import pytest
from sqlalchemy import create_engine, select
from sqlalchemy.orm import Session, selectinload
from typer.testing import CliRunner

from policy_extractor.cli import app
from policy_extractor.export import export_csv, export_xlsx
from policy_extractor.storage.models import Asegurado, Base, Cobertura, Poliza

cli_runner = CliRunner()


# ---------------------------------------------------------------------------
# Fixture helpers
# ---------------------------------------------------------------------------


def _make_poliza(session) -> Poliza:
    """Create a fully-populated Poliza with one Asegurado and one Cobertura.

    Returns the Poliza reloaded with relationships eagerly loaded.
    """
    poliza = Poliza(
        numero_poliza="POL-001",
        aseguradora="AXA",
        tipo_seguro="auto",
        fecha_emision=date(2024, 3, 15),
        inicio_vigencia=date(2024, 4, 1),
        fin_vigencia=date(2025, 4, 1),
        nombre_contratante="Juan Perez",
        nombre_agente="Maria Lopez",
        prima_total=Decimal("15250.50"),
        moneda="MXN",
        forma_pago="anual",
        frecuencia_pago="anual",
        campos_adicionales={
            "confianza": {"overall": 0.95},
            "deducible_umas": 50,
            "copago": 10,
        },
    )
    asegurado = Asegurado(
        tipo="persona",
        nombre_descripcion="Juan Perez",
        fecha_nacimiento=date(1990, 5, 20),
        rfc="PEPJ900520XXX",
        campos_adicionales={"tipo_bien": "vehiculo"},
    )
    cobertura = Cobertura(
        nombre_cobertura="Responsabilidad Civil",
        suma_asegurada=Decimal("500000.00"),
        deducible=Decimal("5000.00"),
        moneda="MXN",
        campos_adicionales={"coaseguro": "20%"},
    )
    poliza.asegurados.append(asegurado)
    poliza.coberturas.append(cobertura)
    session.add(poliza)
    session.commit()

    # Reload with relationships eagerly loaded
    result = session.execute(
        select(Poliza)
        .where(Poliza.id == poliza.id)
        .options(
            selectinload(Poliza.asegurados),
            selectinload(Poliza.coberturas),
        )
    )
    return result.scalar_one()


# ---------------------------------------------------------------------------
# EXP-01: Excel workbook structure tests
# ---------------------------------------------------------------------------


def test_xlsx_sheet_names(session, tmp_path):
    """export_xlsx creates workbook with sheets named polizas, asegurados, coberturas."""
    import openpyxl

    poliza = _make_poliza(session)
    out = tmp_path / "out.xlsx"
    export_xlsx([poliza], out)

    wb = openpyxl.load_workbook(out)
    assert wb.sheetnames == ["polizas", "asegurados", "coberturas"]


def test_xlsx_polizas_headers(session, tmp_path):
    """Polizas sheet row 1 contains all required column headers."""
    import openpyxl

    poliza = _make_poliza(session)
    out = tmp_path / "out.xlsx"
    export_xlsx([poliza], out)

    wb = openpyxl.load_workbook(out)
    ws = wb["polizas"]
    headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]

    required = [
        "numero_poliza", "aseguradora", "tipo_seguro",
        "fecha_emision", "inicio_vigencia", "fin_vigencia",
        "nombre_contratante", "nombre_agente",
        "prima_total", "moneda", "forma_pago", "frecuencia_pago",
    ]
    for col in required:
        assert col in headers, f"Missing column: {col}"


def test_xlsx_asegurados_has_numero_poliza(session, tmp_path):
    """Asegurados sheet has numero_poliza column for VLOOKUP."""
    import openpyxl

    poliza = _make_poliza(session)
    out = tmp_path / "out.xlsx"
    export_xlsx([poliza], out)

    wb = openpyxl.load_workbook(out)
    ws = wb["asegurados"]
    headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
    assert "numero_poliza" in headers


def test_xlsx_coberturas_has_numero_poliza(session, tmp_path):
    """Coberturas sheet has numero_poliza column for VLOOKUP."""
    import openpyxl

    poliza = _make_poliza(session)
    out = tmp_path / "out.xlsx"
    export_xlsx([poliza], out)

    wb = openpyxl.load_workbook(out)
    ws = wb["coberturas"]
    headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
    assert "numero_poliza" in headers


def test_xlsx_campos_expansion(session, tmp_path):
    """campos_adicionales keys appear as individual columns after fixed columns."""
    import openpyxl

    poliza = _make_poliza(session)
    out = tmp_path / "out.xlsx"
    export_xlsx([poliza], out)

    wb = openpyxl.load_workbook(out)
    ws = wb["polizas"]
    headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]

    # campos_adicionales has deducible_umas and copago (confianza should be stripped)
    assert "deducible_umas" in headers
    assert "copago" in headers


# ---------------------------------------------------------------------------
# EXP-02: CSV tests
# ---------------------------------------------------------------------------


def test_csv_writes_polizas_only(session, tmp_path):
    """CSV output has poliza rows — not asegurado/cobertura rows."""
    poliza = _make_poliza(session)
    out = tmp_path / "out.csv"
    export_csv([poliza], out)

    with open(out, encoding="utf-8-sig", newline="") as f:
        reader = csv.DictReader(f)
        rows = list(reader)

    # Should have exactly 1 row (one poliza)
    assert len(rows) == 1
    # Row should have poliza fields
    assert rows[0]["numero_poliza"] == "POL-001"
    # Should NOT have asegurado-specific fields in header
    fieldnames = reader.fieldnames or []
    assert "nombre_descripcion" not in fieldnames


def test_csv_utf8_bom(session, tmp_path):
    """First 3 bytes of output file are UTF-8 BOM (EF BB BF)."""
    poliza = _make_poliza(session)
    out = tmp_path / "out.csv"
    export_csv([poliza], out)

    raw = out.read_bytes()
    assert raw[:3] == b"\xef\xbb\xbf", "CSV missing UTF-8 BOM"


def test_csv_campos_expansion(session, tmp_path):
    """campos_adicionales keys appear as CSV columns."""
    poliza = _make_poliza(session)
    out = tmp_path / "out.csv"
    export_csv([poliza], out)

    with open(out, encoding="utf-8-sig", newline="") as f:
        reader = csv.DictReader(f)
        rows = list(reader)
        fieldnames = reader.fieldnames or []

    assert "deducible_umas" in fieldnames
    assert "copago" in fieldnames
    assert rows[0]["deducible_umas"] == "50"
    assert rows[0]["copago"] == "10"


# ---------------------------------------------------------------------------
# EXP-05: Cell type tests
# ---------------------------------------------------------------------------


def test_xlsx_prima_is_numeric(session, tmp_path):
    """prima_total cell has data_type 'n' (numeric), not 's' (string)."""
    import openpyxl

    poliza = _make_poliza(session)
    out = tmp_path / "out.xlsx"
    export_xlsx([poliza], out)

    wb = openpyxl.load_workbook(out)
    ws = wb["polizas"]
    headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
    prima_col_idx = headers.index("prima_total") + 1

    data_row = list(ws.iter_rows(min_row=2, max_row=2))[0]
    prima_cell = data_row[prima_col_idx - 1]
    assert prima_cell.data_type == "n", (
        f"prima_total data_type is '{prima_cell.data_type}', expected 'n'"
    )


def test_xlsx_date_is_date_type(session, tmp_path):
    """fecha_emision cell is an Excel date — readable as datetime.date by openpyxl."""
    import datetime

    import openpyxl

    poliza = _make_poliza(session)
    out = tmp_path / "out.xlsx"
    export_xlsx([poliza], out)

    wb = openpyxl.load_workbook(out)
    ws = wb["polizas"]
    headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
    fecha_col_idx = headers.index("fecha_emision") + 1

    data_row = list(ws.iter_rows(min_row=2, max_row=2))[0]
    fecha_cell = data_row[fecha_col_idx - 1]
    assert isinstance(fecha_cell.value, (datetime.date, datetime.datetime)), (
        f"fecha_emision value is {type(fecha_cell.value)}, expected date/datetime"
    )


def test_xlsx_date_format(session, tmp_path):
    """fecha_emision cell has number_format == 'DD/MM/YYYY'."""
    import openpyxl

    poliza = _make_poliza(session)
    out = tmp_path / "out.xlsx"
    export_xlsx([poliza], out)

    wb = openpyxl.load_workbook(out)
    ws = wb["polizas"]
    headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
    fecha_col_idx = headers.index("fecha_emision") + 1

    data_row = list(ws.iter_rows(min_row=2, max_row=2))[0]
    fecha_cell = data_row[fecha_col_idx - 1]
    assert fecha_cell.number_format == "DD/MM/YYYY", (
        f"fecha_emision number_format is '{fecha_cell.number_format}', expected 'DD/MM/YYYY'"
    )


def test_xlsx_decimal_converted(session, tmp_path):
    """Decimal('15250.00') written as float 15250.0, not a string."""
    import openpyxl

    poliza = _make_poliza(session)
    out = tmp_path / "out.xlsx"
    export_xlsx([poliza], out)

    wb = openpyxl.load_workbook(out)
    ws = wb["polizas"]
    headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
    prima_col_idx = headers.index("prima_total") + 1

    data_row = list(ws.iter_rows(min_row=2, max_row=2))[0]
    prima_cell = data_row[prima_col_idx - 1]
    assert isinstance(prima_cell.value, float), (
        f"prima_total value is {type(prima_cell.value)}, expected float"
    )
    assert prima_cell.value == pytest.approx(15250.50)


# ---------------------------------------------------------------------------
# Edge cases
# ---------------------------------------------------------------------------


def test_xlsx_empty_polizas(tmp_path):
    """export_xlsx with empty list writes header-only sheets without error."""
    import openpyxl

    out = tmp_path / "out.xlsx"
    export_xlsx([], out)

    wb = openpyxl.load_workbook(out)
    assert wb.sheetnames == ["polizas", "asegurados", "coberturas"]
    ws = wb["polizas"]
    # Should have header row only
    assert ws.max_row == 1


def test_xlsx_confianza_stripped(session, tmp_path):
    """confianza key in campos_adicionales is NOT in column headers."""
    import openpyxl

    poliza = _make_poliza(session)
    out = tmp_path / "out.xlsx"
    export_xlsx([poliza], out)

    wb = openpyxl.load_workbook(out)
    ws = wb["polizas"]
    headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
    assert "confianza" not in headers


def test_csv_comma_in_value(session, tmp_path):
    """Field containing comma is properly quoted in CSV output."""
    # Create a poliza with a comma-containing value
    poliza = Poliza(
        numero_poliza="POL-002",
        aseguradora="Qualitas, S.A.",  # comma in value
        moneda="MXN",
    )
    session.add(poliza)
    session.commit()

    from sqlalchemy import select
    from sqlalchemy.orm import selectinload

    result = session.execute(
        select(Poliza)
        .where(Poliza.id == poliza.id)
        .options(
            selectinload(Poliza.asegurados),
            selectinload(Poliza.coberturas),
        )
    )
    poliza = result.scalar_one()

    out = tmp_path / "out.csv"
    export_csv([poliza], out)

    # Raw bytes should contain properly quoted value
    content = out.read_text(encoding="utf-8-sig")
    # The csv module should quote "Qualitas, S.A." in the output
    assert '"Qualitas, S.A."' in content


# ---------------------------------------------------------------------------
# CLI integration tests — EXP-03 (format flags), EXP-04 (filter parity)
# ---------------------------------------------------------------------------


def _make_cli_engine_and_factory():
    """Create an in-memory engine and a mock SessionLocal factory for CLI tests."""
    from sqlalchemy.orm import sessionmaker

    engine = create_engine("sqlite:///:memory:")
    Base.metadata.create_all(engine)

    RealSession = sessionmaker(bind=engine)

    class _FakeSessionCls:
        def __call__(self):
            return RealSession()

        def configure(self, **kwargs):
            pass

    return engine, _FakeSessionCls()


def _insert_poliza(engine, **kwargs) -> None:
    """Insert a Poliza into engine using provided keyword args."""
    poliza = Poliza(**kwargs)
    with Session(engine) as s:
        s.add(poliza)
        s.commit()


def test_cli_export_xlsx(tmp_path):
    """CLI export --format xlsx -o out.xlsx produces a valid xlsx with 3 sheets."""
    import openpyxl

    engine, factory = _make_cli_engine_and_factory()
    _insert_poliza(
        engine,
        numero_poliza="POL-001",
        aseguradora="AXA",
        tipo_seguro="auto",
        inicio_vigencia=date(2024, 4, 1),
        fin_vigencia=date(2025, 4, 1),
        prima_total=Decimal("1000.00"),
        moneda="MXN",
    )

    out = tmp_path / "out.xlsx"
    with (
        patch("policy_extractor.cli.init_db"),
        patch("policy_extractor.cli.SessionLocal", factory),
    ):
        result = cli_runner.invoke(app, ["export", "--format", "xlsx", "-o", str(out)])

    assert result.exit_code == 0, result.output
    assert out.exists()
    wb = openpyxl.load_workbook(out)
    assert wb.sheetnames == ["polizas", "asegurados", "coberturas"]


def test_cli_export_csv(tmp_path):
    """CLI export --format csv -o out.csv produces a valid UTF-8 BOM CSV with header."""
    engine, factory = _make_cli_engine_and_factory()
    _insert_poliza(
        engine,
        numero_poliza="POL-001",
        aseguradora="AXA",
        moneda="MXN",
    )

    out = tmp_path / "out.csv"
    with (
        patch("policy_extractor.cli.init_db"),
        patch("policy_extractor.cli.SessionLocal", factory),
    ):
        result = cli_runner.invoke(app, ["export", "--format", "csv", "-o", str(out)])

    assert result.exit_code == 0, result.output
    assert out.exists()
    # Must have UTF-8 BOM
    raw = out.read_bytes()
    assert raw[:3] == b"\xef\xbb\xbf", "CSV missing UTF-8 BOM"
    # Must have header row
    content = out.read_text(encoding="utf-8-sig")
    assert "numero_poliza" in content


def test_cli_export_json_default(tmp_path):
    """CLI export with no --format outputs valid JSON array to stdout (backward compat)."""
    engine, factory = _make_cli_engine_and_factory()
    _insert_poliza(engine, numero_poliza="POL-001", aseguradora="AXA")

    with (
        patch("policy_extractor.cli.init_db"),
        patch("policy_extractor.cli.SessionLocal", factory),
    ):
        result = cli_runner.invoke(app, ["export"])

    assert result.exit_code == 0, result.output
    data = json.loads(result.output)
    assert isinstance(data, list)
    assert len(data) == 1
    assert data[0]["numero_poliza"] == "POL-001"


def test_cli_export_xlsx_requires_output():
    """CLI export --format xlsx without -o exits 1 with 'required' in output."""
    engine, factory = _make_cli_engine_and_factory()

    with (
        patch("policy_extractor.cli.init_db"),
        patch("policy_extractor.cli.SessionLocal", factory),
    ):
        result = cli_runner.invoke(app, ["export", "--format", "xlsx"])

    assert result.exit_code == 1
    assert "required" in result.output.lower()


def test_xlsx_filter_aseguradora(tmp_path):
    """--aseguradora AXA exports only AXA polizas; GNP poliza excluded."""
    import openpyxl

    engine, factory = _make_cli_engine_and_factory()
    _insert_poliza(engine, numero_poliza="POL-AXA", aseguradora="AXA", moneda="MXN")
    _insert_poliza(engine, numero_poliza="POL-GNP", aseguradora="GNP", moneda="MXN")

    out = tmp_path / "out.xlsx"
    with (
        patch("policy_extractor.cli.init_db"),
        patch("policy_extractor.cli.SessionLocal", factory),
    ):
        result = cli_runner.invoke(
            app,
            ["export", "--aseguradora", "AXA", "--format", "xlsx", "-o", str(out)],
        )

    assert result.exit_code == 0, result.output
    assert out.exists()
    wb = openpyxl.load_workbook(out)
    ws = wb["polizas"]
    # Row 1 is header; row 2 is the single AXA poliza
    assert ws.max_row == 2, f"Expected 2 rows (header + 1 data), got {ws.max_row}"
    # Verify the data row is AXA
    headers = [cell.value for cell in list(ws.iter_rows(min_row=1, max_row=1))[0]]
    aseg_col = headers.index("aseguradora") + 1
    data_row = list(ws.iter_rows(min_row=2, max_row=2))[0]
    assert data_row[aseg_col - 1].value == "AXA"


def test_xlsx_filter_dates(tmp_path):
    """--desde 2024-06-01 exports only polizas with inicio_vigencia >= 2024-06-01."""
    import openpyxl

    engine, factory = _make_cli_engine_and_factory()
    # Before cutoff
    _insert_poliza(
        engine,
        numero_poliza="POL-OLD",
        aseguradora="AXA",
        inicio_vigencia=date(2024, 1, 1),
        moneda="MXN",
    )
    # After cutoff
    _insert_poliza(
        engine,
        numero_poliza="POL-NEW",
        aseguradora="AXA",
        inicio_vigencia=date(2024, 7, 1),
        moneda="MXN",
    )

    out = tmp_path / "out.xlsx"
    with (
        patch("policy_extractor.cli.init_db"),
        patch("policy_extractor.cli.SessionLocal", factory),
    ):
        result = cli_runner.invoke(
            app,
            ["export", "--desde", "2024-06-01", "--format", "xlsx", "-o", str(out)],
        )

    assert result.exit_code == 0, result.output
    assert out.exists()
    wb = openpyxl.load_workbook(out)
    ws = wb["polizas"]
    # Row 1 header + 1 data row (POL-NEW only)
    assert ws.max_row == 2, f"Expected 2 rows (header + 1 data), got {ws.max_row}"
    headers = [cell.value for cell in list(ws.iter_rows(min_row=1, max_row=1))[0]]
    pol_col = headers.index("numero_poliza") + 1
    data_row = list(ws.iter_rows(min_row=2, max_row=2))[0]
    assert data_row[pol_col - 1].value == "POL-NEW"
